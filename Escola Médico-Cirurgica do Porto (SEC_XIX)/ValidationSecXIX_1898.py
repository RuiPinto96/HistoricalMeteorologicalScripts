import unicodedata
import sys
import time
import datetime
import calendar
from influxdb import InfluxDBClient
from ValidationValueLimits import temp_max, temp_min, pressure_max, pressure_min, vapor_max, vapor_min, humidity_max, humidity_min, ozone_max, ozone_min, cloud_max, cloud_min, pluv_max, pluv_min, udo_max, udo_min, abso_wind_speed_max, abso_wind_speed_min, clock_wind_speed_max, clock_wind_speed_min

from os import listdir
from os.path import isfile, join
onlyfiles = [f for f in listdir('1898') if isfile(join('1898', f))]

from openpyxl import load_workbook

print(onlyfiles)


for file in onlyfiles:
    print('')
    print('-----------------------------------------------------')
    print('|       ' + '                                            |')
    print('       ' + file + '       ')
    print('|       ' + '                                            |')
    
    print('-----------------------------------------------------')
    file='1898/' + file
    wb=load_workbook(filename=str(file), data_only=True)



    number_sheets= len(wb.sheetnames)

    #print(number_sheets)

    user = ''
    password = ''
    host = 'localhost'
    dbname = 'Dados_Sec_XIX_igup'
    

    for index3,ws in enumerate(wb.worksheets):
        if(index3 < 12):
            if(ws==wb.worksheets[0]):
                cellyear = ws.cell(row=2, column=13).value
                if(cellyear == None):
                    cellyear = ws.cell(row=4, column=13).value
                newcellyear = cellyear[-4:]
                year = newcellyear
                ano_ver=int(year)
            
            temp = ws.cell(row=2, column=13).value
            if(temp == None):
                temp = ws.cell(row=4, column=13).value
            temp1 = temp[-4:]
            ano = int(temp1)
            
        #FEVEREIRO
        if ws==wb.worksheets[1]:
            if(calendar.isleap(ano_ver)):
                minr=20
                maxc=32
                maxr=48
                minc=2
            else:
                minr=20
                maxc=32
                maxr=47
                minc=2
                
                
        #Mes com 31 dias
        if ws==wb.worksheets[0] or ws==wb.worksheets[2]:
            if(ws.cell(row=2, column=13).value == None):
                minr=22
                maxc=32
                maxr=52
                minc=2
            else:
                minr=20
                maxc=32
                maxr=50
                minc=2
           
           
       
            
        if index3 <= 2:
            for index2,row in enumerate(ws.iter_rows(min_row=minr, max_col=maxc, max_row=maxr, min_col=minc, values_only=True)):
                for index, cell in enumerate(row):
                   
                    if(index3!=0):
                        anoValidation = ws.cell(row=2, column=13).value
                        if(anoValidation == None):
                            anoValidation = ws.cell(row=4, column=13).value
                        newAnoValidation = anoValidation[-4:]
                        AnoVali = int(newAnoValidation)
                        if (ano!= AnoVali):
                            print('Ano esta mal! ' + file)
                    dia=index2 + 1
                    
                    if(index3==0):
                        mes=1
                    if(index3==1):
                        mes = 2
                    if(index3==2):
                        mes = 3 
                    
                    
                    
                    if(index<31 and index>=0):
                        if(index==1 or index==5 or index==6 or index==14 or index==17):
                            hora = 9
                            try:
                                value=float(cell)
                                
                                #Valores Temperatura fora do normal
                                if(index==5 or index==6):
                                    if(value < temp_max or value >= temp_min):
                                        value_temp = value
                                
                                #Valores Barometro fora do normal
                                if(index==1):
                                    if(value < pressure_max or value >= pressure_min):
                                        value_barometer = value

                                #Valores Tensao Vapor fora do normal
                                if(index==14):
                                    if(value < vapor_max or value >= vapor_min):
                                        value_vapor = value
                                        
                                #Valores Humidade fora do normal
                                if(index==17):
                                    if(value < humidity_max or value >= humidity_min):
                                        value_humidity = value
                                
                            except ValueError:
                                print('Formato de Valor incorrecto: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora) + ' Coluna ' + str(index))
                            except TypeError:
                                print('Celula vazia: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora ' + str(hora) + ' Coluna ' + str(index))
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia,hora)
                            #formated_timestamp = (ano) + ',' + str(mes) + ',' + str(dia) + ',' + str(hora)
                            formated_timestamp = str(ano) + '-' + str(mes) + '-' + str(dia) + 'T' + str(hora)+ ':00:00Z'
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d,%H')
                            #ut = time.mktime(dt.timetuple())
                            #print ut
                        if(index==2 or index==7 or index==8 or index==15 or index==18 or index==21  or index==27):
                            hora = 12
                            try:
                                value=float(cell)
                                
                                #Temperatura fora do normal
                                if(index==7 or index==8):
                                    if(value < temp_max or value >= temp_min):
                                        value_temp = value
                                
                                #Valores Barometro fora do normal
                                if(index==2):
                                    if(value < pressure_max or value >= pressure_min):
                                        value_barometer = value
                                
                                #Valores Tensao Vapor fora do normal
                                if(index==15):
                                    if(value < vapor_max or value >= vapor_min):
                                        value_vapor = value
                                        
                                #Valores Humidade fora do normal
                                if(index==18):
                                    if(value < humidity_max or value >= humidity_min):
                                        value_humidity = value

                                #Valores Ozonometro fora do normal
                                if(index==21):
                                    if(value < ozone_max or value >= ozone_min):
                                       value_ozone = value

                                #Valores Quantidade de Nuvens fora do normal
                                if(index==27):
                                    if(value > cloud_max or value < cloud_min):
                                        print('Valor Suspeito de Quantidade de Nuvens (12H): ' + str(value) + ' - ' + str(ws) + ' Dia: ' + str(dia))
                                
                            except ValueError:
                                print('Formato de Valor incorrecto: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora) + ' Coluna ' + str(index))
                            except TypeError:
                                print('Celula vazia: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora ' + str(hora) + ' Coluna ' + str(index))
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia,hora)
                            #formated_timestamp = (ano) + ',' + str(mes) + ',' + str(dia) + ',' + str(hora)
                            formated_timestamp = str(ano) + '-' + str(mes) + '-' + str(dia) + 'T' + str(hora)+ ':00:00Z'
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d,%H')
                            #ut = time.mktime(dt.timetuple())
                            #print ut

                        
                        if(index==3 or index==9 or index==10 or index==11 or index==12 or index==16 or index==19 or index==20):
                            hora = 15
                            try:
                                value=float(cell)
                                
                                #Temperatura fora do normal
                                if(index==9 or index==10 or index==11 or index==12):
                                    if(value < temp_max or value >= temp_min):
                                        value_temp = value
                                
                                #Valores Barometro fora do normal
                                if(index==3):
                                    if(value < pressure_max or value >= pressure_min):
                                        value_barometer = value
                                        
                                #Valores Tensao Vapor fora do normal
                                if(index==16):
                                    if(value < vapor_max or value >= vapor_min):
                                        value_vapor = value
                                    
                                #Valores Humidade fora do normal
                                if(index==19):
                                    if(value < humidity_max or value >= humidity_min):
                                        value_humidity = value
                                        
                                #Valores Pluvimetro fora do normal
                                if(index==20):
                                    if(value < pluv_max or value >= pluv_min):
                                        value_pluv = value
                                
                            except ValueError:
                                print('Formato de Valor incorrecto: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora) + ' Coluna ' + str(index))
                            except TypeError:
                                print('Celula vazia: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora ' + str(hora) + ' Coluna ' + str(index))
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia,hora)
                            #formated_timestamp = (ano) + ',' + str(mes) + ',' + str(dia) + ',' + str(hora)
                            formated_timestamp = str(ano) + '-' + str(mes) + '-' + str(dia) + 'T' + str(hora)+ ':00:00Z'
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d,%H')
                            #ut = time.mktime(dt.timetuple())
                            #print ut
                            
    
                        if(index==4 or index==13):
                            try:
                                value=float(cell)
                                
                                #Temperatura fora do normal
                                if(index==13):
                                    if(value >= temp_max or value < temp_min):
                                        print('Valor Suspeito Temperatura (Media): ' + str(value) + ' - ' + str(ws) + ' Dia: ' + str(dia))
                                
                                #Valores Barometro fora do normal
                                if(index==4):
                                    if(value >=pressure_max or value < pressure_min):
                                        print('Valor Suspeito Pressao (Media): ' + str(value) + ' - ' + str(ws) + ' Dia: ' + str(dia))
                                
                            except ValueError:
                                print('Formato de Valor incorrecto: ' + str(ws) + ' Dia: ' + str(dia) + ' Coluna ' + str(index))
                            except TypeError:
                                print('Celula vazia: ' + str(ws) + ' Dia: ' + str(dia) + ' Coluna ' + str(index))
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia)
                            #formated_timestamp = str(ano) + ',' + str(mes) + ',' + str(dia)
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d')
                            #ut = time.mktime(dt.timetuple())
                            #print ut

                        if(index==22 or index==28): 
                            hora = 9
                            try: 
                                value = unicodedata.normalize('NFD', cell).encode('ascii', 'ignore')
                            except ValueError:
                                print('Formato de Valor incorrecto: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora) + ' Coluna ' + str(index))
                            except TypeError:
                                print('Celula vazia: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora ' + str(hora) + ' Coluna ' + str(index))
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia)
                            #formated_timestamp = str(ano) + ',' + str(mes) + ',' + str(dia)
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d')
                            #ut = time.mktime(dt.timetuple())
                            #print ut


                        if(index==23 or index==29):
                            hora = 12
                            try: 
                                value = unicodedata.normalize('NFD', cell).encode('ascii', 'ignore')
                            except ValueError:
                                print('Formato de Valor incorrecto: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora) + ' Coluna ' + str(index))
                            except TypeError:
                                print('Celula vazia: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora ' + str(hora) + ' Coluna ' + str(index))
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia)
                            #formated_timestamp = str(ano) + ',' + str(mes) + ',' + str(dia)
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d')
                            #ut = time.mktime(dt.timetuple())
                            #print ut

                        if(index==24 or index==30):
                            hora = 15
                            try: 
                                value = unicodedata.normalize('NFD', cell).encode('ascii', 'ignore')
                            except ValueError:
                                print('Formato de Valor incorrecto: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora) + ' Coluna ' + str(index))
                            except TypeError:
                                print('Celula vazia: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora ' + str(hora) + ' Coluna ' + str(index))
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia)
                            #formated_timestamp = str(ano) + ',' + str(mes) + ',' + str(dia)
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d')
                            #ut = time.mktime(dt.timetuple())
                            #print ut
                            
                        if(index==25 or index==26):
                            horaAnemometro = ws.cell(row=16, column=27).value
                            if(horaAnemometro!= 'Meio dia'):
                                hora = 15
                            if(horaAnemometro == 'Meio dia'):
                                hora = 12
                            try:
                                value=float(cell)
                                
                                #Valores Velocidade Absoluta fora do normal km/h
                                if(index==25):
                                    if(value > abso_wind_speed_max or value < abso_wind_speed_min):
                                        print('Valor Suspeito Velocidade Absoluta  do Vento (' + str(hora) + 'H): '  + str(value) + ' - ' + str(ws) + ' Dia: ' + str(dia))
                                
                                #Valores Velocidade Horaria fora do normal
                                if(index==26):
                                    if(value > clock_wind_speed_max or value < clock_wind_speed_min):
                                        print('Valor Suspeito Velocidade Horaria  do Vento (' + str(hora) + 'H): '  + str(value) + ' - ' + str(ws) + ' Dia: ' + str(dia))
                                
                            except ValueError:
                                print('Formato de Valor incorrecto: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora) + ' Coluna ' + str(index))
                            except TypeError:
                                print('Celula vazia: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora ' + str(hora) + ' Coluna ' + str(index))
                                pass
                            except AttributeError:
                                pass
                            #print(cell,int(ano),mes,dia,hora)
                            #formated_timestamp = (ano) + ',' + str(mes) + ',' + str(dia) + ',' + str(hora)
                            #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d,%H')
                            #ut = time.mktime(dt.timetuple())
                            #print ut
                            
                json_body = [
                                {
                                    "measurement": "SEC_XIX_meteorological_data",
                                    "tags": {
                                        "location": "Escola medico-cirurgica do Porto.",
                                        "latitude": 41.13861111,
                                        "longitude": 8.60250000,
                                        "altitude": 84.795,
                                    },
                                    "time": formated_timestamp,
                                    "fields": {
                                        "Barometer_mm": value_barometer,
                                        "Temperature_C": value_temp,
                                        "Vapor_Tension_mm": value_vapor,
                                        "Humidity_%": value_humidity,
                                        "Ozone_grains": value_ozone,
                                        "Pluvimeter_mm": value_pluv
                                    }
                                }
                            ]
                                        
                client = InfluxDBClient(host,8086,user,password,dbname)

                client.write_points(json_body)
                

    print('--------------------------------------------------------------------------------------------') 
    print('--------------------------------------------------------------------------------------------') 
                            
